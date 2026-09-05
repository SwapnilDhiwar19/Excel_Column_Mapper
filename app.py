import io
import json
import re
from collections import Counter

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Direct Databricks native connector
try:
    from databricks import sql as dbsql
    HAS_DATABRICKS_DRIVER = True
except ImportError:
    HAS_DATABRICKS_DRIVER = False

# Fuzzy matching engine setup
try:
    from rapidfuzz import fuzz
    FUZZ_BACKEND = "rapidfuzz"
except ImportError:
    from difflib import SequenceMatcher
    FUZZ_BACKEND = "difflib"

    class _FuzzShim:
        @staticmethod
        def token_sort_ratio(a: str, b: str) -> float:
            return SequenceMatcher(None, a, b).ratio() * 100

    fuzz = _FuzzShim()


# ============================================================================
# Core Normalization & Matching Logic
# ============================================================================

def normalize(text_val: str) -> str:
    """Handle CamelCase, PascalCase, separators, and collapse whitespace."""
    text_val = str(text_val).strip()
    text_val = re.sub(r"([a-z])([A-Z])", r"\1 \2", text_val)
    text_val = re.sub(r"([A-Z]+)([A-Z][a-z])", r"\1 \2", text_val)
    text_val = text_val.lower()
    text_val = re.sub(r"[_\-]+", " ", text_val)
    text_val = re.sub(r"\s+", " ", text_val)
    return text_val.strip()


def build_targets(db_columns: list[dict]) -> list[dict]:
    targets = []
    for entry in db_columns:
        db_column = str(entry.get("db_column", "")).strip()
        if not db_column:
            continue
        aliases = [str(a).strip() for a in entry.get("aliases", []) if str(a).strip()]
        targets.append({
            "db_column": db_column,
            "aliases": aliases,
            "norm_db": normalize(db_column),
            "norm_aliases": [normalize(a) for a in aliases],
        })
    return targets


def match_column(excel_col: str, targets: list[dict], threshold: float) -> dict:
    norm_col = normalize(excel_col)

    # 1. Exact DB match
    for t in targets:
        if norm_col == t["norm_db"]:
            return {"db_column": t["db_column"], "match_type": "Exact", "score": 100.0}

    # 2. Exact Alias match
    for t in targets:
        if norm_col in t["norm_aliases"]:
            return {"db_column": t["db_column"], "match_type": "Alias", "score": 100.0}

    # 3. Fuzzy match
    best = {"db_column": None, "match_type": "Unmatched", "score": 0.0}
    for t in targets:
        for cand in [t["norm_db"]] + t["norm_aliases"]:
            score = fuzz.token_sort_ratio(norm_col, cand)
            if score > best["score"]:
                best = {"db_column": t["db_column"], "match_type": "Fuzzy", "score": float(score)}

    if best["score"] >= threshold:
        return best
    return {"db_column": None, "match_type": "Unmatched", "score": best["score"]}


def map_columns(excel_columns: list[str], targets: list[dict], threshold: float) -> list[dict]:
    raw_results = [
        {"excel_column": col, **match_column(col, targets, threshold)}
        for col in excel_columns
    ]

    claimed = {}
    for r in raw_results:
        if r["db_column"] is None:
            continue
        key = r["db_column"]
        if key not in claimed or r["score"] > claimed[key]["score"]:
            claimed[key] = r
    winners = {id(r) for r in claimed.values()}

    final_results = []
    for r in raw_results:
        if r["db_column"] is not None and id(r) not in winners:
            final_results.append({
                "excel_column": r["excel_column"],
                "db_column": None,
                "match_type": "Unmatched (duplicate)",
                "score": r["score"],
            })
        else:
            final_results.append(r)
    return final_results


def build_output_workbook_bytes(
    mapped_df: pd.DataFrame,
    mapping_results: list[dict],
    restrict_report_to: list[str] | None = None,
) -> bytes:
    report_rows = mapping_results
    if restrict_report_to is not None:
        allowed = set(restrict_report_to)
        report_rows = [
            r for r in mapping_results
            if (r.get("db_column") or r["excel_column"]) in allowed
        ]

    report_df = pd.DataFrame([
        {
            "Excel Column (original)": r["excel_column"],
            "Mapped DB Column": r.get("db_column") if r.get("db_column") else "-- NOT MATCHED --",
            "Match Type": r.get("match_type", "Manual"),
            "Confidence Score": round(float(r.get("score", 0.0)), 1),
        }
        for r in report_rows
    ])

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        mapped_df.to_excel(writer, sheet_name="Mapped Data", index=False)
        report_df.to_excel(writer, sheet_name="Column Mapping Report", index=False)
    buffer.seek(0)

    return _style_report_sheet(buffer.read(), len(report_df))


def _style_report_sheet(workbook_bytes: bytes, n_rows: int) -> bytes:
    wb = load_workbook(io.BytesIO(workbook_bytes))
    ws = wb["Column Mapping Report"]

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    unmatched_fill = PatternFill(start_color="FFE4E6", end_color="FFE4E6", fill_type="solid")
    fuzzy_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    exact_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx in range(2, n_rows + 2):
        match_type = ws.cell(row=row_idx, column=3).value
        if match_type and ("Unmatched" in str(match_type) or "Duplicate" in str(match_type)):
            fill = unmatched_fill
        elif match_type == "Fuzzy":
            fill = fuzzy_fill
        else:
            fill = exact_fill
        for cell in ws[row_idx]:
            cell.fill = fill

    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 14), 50)
    ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _badge(match_type: str) -> str:
    if not match_type or "Unmatched" in match_type:
        return f"❌ {match_type or 'Unmatched'}"
    if match_type == "Manual":
        return "✏️ Manual Override"
    if match_type == "Fuzzy":
        return "⚡ Fuzzy Match"
    if match_type == "Alias":
        return "🏷️ Known Alias"
    return "✅ Exact Match"


def load_uploaded_df(file) -> pd.DataFrame:
    if file.name.endswith(".csv"):
        return pd.read_csv(file)
    xls = pd.ExcelFile(file)
    return pd.read_excel(file, sheet_name=xls.sheet_names[0])


# ============================================================================
# Databricks Native Driver Helpers (Direct Secrets Integration)
# ============================================================================

def get_databricks_connection():
    if not HAS_DATABRICKS_DRIVER:
        st.error("Missing dependency: Run `pip install databricks-sql-connector` in your environment.")
        st.stop()

    if "databricks" not in st.secrets:
        st.error("Missing Databricks configuration: Add `[databricks]` section to your Streamlit secrets.")
        st.stop()

    sec = st.secrets["databricks"]
    clean_host = sec.get("host", "").replace("https://", "").replace("http://", "").strip("/")
    clean_path = sec.get("http_path", "").strip()
    if clean_path and not clean_path.startswith("/"):
        clean_path = "/" + clean_path

    clean_token = sec.get("token", "").strip().strip("'").strip('"')
    if clean_token.lower().startswith("bearer "):
        clean_token = clean_token[7:].strip()

    catalog = sec.get("catalog", "workspace").strip()
    schema = sec.get("schema", "excel_column_mapping_utility").strip()

    conn_kwargs = {
        "server_hostname": clean_host,
        "http_path": clean_path,
        "access_token": clean_token,
    }
    if catalog:
        conn_kwargs["catalog"] = catalog
    if schema:
        conn_kwargs["schema"] = schema

    return dbsql.connect(**conn_kwargs)


def format_table_identifier(tbl: str) -> str:
    sec = st.secrets.get("databricks", {})
    cat = sec.get("catalog", "workspace").strip()
    sch = sec.get("schema", "excel_column_mapping_utility").strip()

    parts = [p.strip().replace("`", "") for p in tbl.strip().split(".") if p.strip()]
    if len(parts) == 3:
        return f"`{parts[0]}`.`{parts[1]}`.`{parts[2]}`"
    elif len(parts) == 2:
        if cat:
            return f"`{cat}`.`{parts[0]}`.`{parts[1]}`"
        return f"`{parts[0]}`.`{parts[1]}`"
    else:
        if cat and sch:
            return f"`{cat}`.`{sch}`.`{parts[0]}`"
        elif sch:
            return f"`{sch}`.`{parts[0]}`"
        return f"`{parts[0]}`"


# ============================================================================
# Streamlit Interface
# ============================================================================

st.set_page_config(
    page_title="AutoSchema Mapper - Databricks",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
    <style>
        .main { background-color: #f8fafc; }
        .hero-container {
            background: linear-gradient(135deg, #1e293b 0%, #0f172a 100%);
            padding: 2.2rem 2.5rem;
            border-radius: 16px;
            color: #ffffff;
            margin-bottom: 2rem;
            box-shadow: 0 10px 25px -5px rgba(15, 23, 42, 0.15);
        }
        .hero-title {
            font-size: 2.2rem;
            font-weight: 700;
            letter-spacing: -0.02em;
            margin-bottom: 0.4rem;
        }
        .hero-desc {
            color: #94a3b8;
            font-size: 1.05rem;
            margin: 0;
            max-width: 850px;
        }
        .metric-card {
            background: #ffffff;
            border: 1px solid #e2e8f0;
            border-radius: 12px;
            padding: 1.25rem 1.5rem;
            box-shadow: 0 2px 4px rgba(0,0,0,0.02);
        }
        .metric-num {
            font-size: 1.85rem;
            font-weight: 700;
            color: #0f172a;
            line-height: 1.2;
        }
        .metric-label {
            font-size: 0.85rem;
            font-weight: 600;
            color: #64748b;
            text-transform: uppercase;
            letter-spacing: 0.05em;
        }
        .step-header {
            font-size: 1.25rem;
            font-weight: 600;
            color: #1e293b;
            margin-bottom: 1rem;
        }
        .col-pill {
            display: inline-block;
            background-color: #e2e8f0;
            color: #0f172a;
            padding: 4px 10px;
            border-radius: 6px;
            font-size: 0.85rem;
            font-family: monospace;
            margin: 3px;
        }
        .pill-mapped {
            display: inline-block;
            background-color: #dcfce7;
            color: #15803d;
            border: 1px solid #bbf7d0;
            padding: 4px 10px;
            border-radius: 6px;
            font-size: 0.85rem;
            font-family: monospace;
            margin: 3px;
            font-weight: 600;
        }
        .pill-unmapped {
            display: inline-block;
            background-color: #ffe4e6;
            color: #b91c1c;
            border: 1px solid #fecdd3;
            padding: 4px 10px;
            border-radius: 6px;
            font-size: 0.85rem;
            font-family: monospace;
            margin: 3px;
            font-weight: 600;
        }
        .pill-master {
            display: inline-block;
            background-color: #e0f2fe;
            color: #0369a1;
            border: 1px solid #bae6fd;
            padding: 4px 10px;
            border-radius: 6px;
            font-size: 0.85rem;
            font-family: monospace;
            margin: 3px;
            font-weight: 600;
        }
    </style>
""", unsafe_allow_html=True)

st.markdown("""
    <div class="hero-container">
        <div class="hero-title">⚡ AutoSchema Mapper (Databricks)</div>
        <p class="hero-desc">Directly connected to Databricks Lakehouse. Select your metadata table, pick your report filter, and map spreadsheet headers with up to 5 master joins.</p>
    </div>
""", unsafe_allow_html=True)

with st.sidebar:
    st.markdown("### ⚙️ Engine Settings")
    threshold = st.slider(
        "Fuzzy Match Sensitivity",
        min_value=50, max_value=100, value=75, step=5,
        format="%d%%",
        help="Higher values require closer spelling matches before auto-accepting."
    )
    st.markdown("---")
    st.markdown("### ℹ️ Databricks Context")
    sec_cfg = st.secrets.get("databricks", {})
    st.caption(f"**Catalog:** `{sec_cfg.get('catalog', 'workspace')}`")
    st.caption(f"**Schema:** `{sec_cfg.get('schema', 'excel_column_mapping_utility')}`")
    st.caption(f"**Fuzzy Engine:** `{FUZZ_BACKEND}`")

# --- 1. Target Schema Configuration (Auto-Connected to Databricks) ---
with st.container():
    st.markdown('<div class="step-header">🎯 1. Target Schema Definition</div>', unsafe_allow_html=True)

    if "databricks_tables" not in st.session_state:
        try:
            with st.spinner("Connecting to Databricks Lakehouse..."):
                conn = get_databricks_connection()
                with conn.cursor() as cursor:
                    cursor.execute("SHOW TABLES")
                    res = cursor.fetchall()
                    found_tables = [row["tableName"] if isinstance(row, dict) else row[1] for row in res]
                    st.session_state["databricks_tables"] = sorted(found_tables)
                conn.close()
        except Exception as e:
            st.error(f"Databricks Connection Failed: {e}")
            st.stop()

    tables_available = st.session_state.get("databricks_tables", [])
    db_columns_input = []

    config_method = st.radio(
        "Extraction Method:",
        ["Filter by Report Name (Cascading Dropdowns)", "Write Custom SQL Query"],
        horizontal=True
    )

    if config_method == "Filter by Report Name (Cascading Dropdowns)":
        c_tbl, c_refresh = st.columns([4, 1])
        with c_tbl:
            if tables_available:
                default_idx = 0
                for idx, t_name in enumerate(tables_available):
                    if "columns_master" in t_name.lower():
                        default_idx = idx
                        break
                selected_table = st.selectbox("1️⃣ Select Metadata Table:", options=tables_available, index=default_idx)
            else:
                selected_table = st.text_input("1️⃣ Metadata Table Name:", value="columns_master")

        with c_refresh:
            st.write("")
            st.write("")
            if st.button("🔄 Refresh Cache"):
                st.session_state.pop(f"cols_{selected_table}", None)
                st.session_state.pop(f"filter_vals_{selected_table}", None)
                st.rerun()

        table_cols_key = f"cols_{selected_table}"
        if table_cols_key not in st.session_state and selected_table:
            try:
                conn = get_databricks_connection()
                qual_tbl = format_table_identifier(selected_table)
                with conn.cursor() as cursor:
                    cursor.execute(f"DESCRIBE TABLE {qual_tbl}")
                    desc_res = cursor.fetchall()
                    loaded_fields = [
                        (row["col_name"] if isinstance(row, dict) else row[0])
                        for row in desc_res
                        if row and not str(row[0]).startswith("#")
                    ]
                    st.session_state[table_cols_key] = loaded_fields
                conn.close()
            except Exception as e:
                st.error(f"Could not read columns from `{selected_table}`: {e}")

        table_fields = st.session_state.get(table_cols_key, [])

        if table_fields:
            col_step2, col_step3 = st.columns(2)

            def find_best_index(options, candidates):
                for c in candidates:
                    for idx, opt in enumerate(options):
                        if c in opt.lower():
                            return idx
                return 0

            with col_step2:
                filter_col_idx = find_best_index(table_fields, ["report", "type", "category"])
                filter_column = st.selectbox("2️⃣ Filter Column (Report Field):", options=table_fields, index=filter_col_idx)

            filter_cache_key = f"filter_vals_{selected_table}_{filter_column}"
            if filter_cache_key not in st.session_state:
                try:
                    conn = get_databricks_connection()
                    qual_tbl = format_table_identifier(selected_table)
                    with conn.cursor() as cursor:
                        cursor.execute(f"SELECT DISTINCT `{filter_column}` FROM {qual_tbl} WHERE `{filter_column}` IS NOT NULL ORDER BY 1")
                        distinct_rows = cursor.fetchall()
                        st.session_state[filter_cache_key] = [str(r[0]) for r in distinct_rows if r[0] is not None]
                    conn.close()
                except Exception as e:
                    st.error(f"Could not fetch distinct values for `{filter_column}`: {e}")

            distinct_filter_values = st.session_state.get(filter_cache_key, [])

            with col_step3:
                if distinct_filter_values:
                    chosen_report = st.selectbox(
                        f"3️⃣ Select Report to Filter on (`{filter_column}`):",
                        options=distinct_filter_values
                    )
                else:
                    chosen_report = None
                    st.warning(f"No values found in `{filter_column}`.")

            if chosen_report:
                try:
                    conn = get_databricks_connection()
                    qual_tbl = format_table_identifier(selected_table)

                    remaining_cols = [c for c in table_fields if c != filter_column]
                    target_candidates = [c for c in remaining_cols if any(k in c.lower() for k in ["column", "target", "col", "field"])]
                    target_col = target_candidates[0] if target_candidates else (remaining_cols[0] if remaining_cols else None)

                    alias_candidates = [c for c in remaining_cols if c != target_col and any(k in c.lower() for k in ["alias", "synonym"])]
                    alias_col = alias_candidates[0] if alias_candidates else None

                    with conn.cursor() as cursor:
                        if target_col and alias_col:
                            query = f"SELECT `{target_col}`, `{alias_col}` FROM {qual_tbl} WHERE `{filter_column}` = '{chosen_report}'"
                            cursor.execute(query)
                            results = cursor.fetchall()
                            db_columns_input = [
                                {
                                    "db_column": str(r[0]).strip(),
                                    "aliases": [a.strip() for a in str(r[1] or "").split(",") if a.strip()]
                                }
                                for r in results if r[0]
                            ]
                        elif target_col:
                            query = f"SELECT `{target_col}` FROM {qual_tbl} WHERE `{filter_column}` = '{chosen_report}'"
                            cursor.execute(query)
                            results = cursor.fetchall()
                            db_columns_input = [{"db_column": str(r[0]).strip(), "aliases": []} for r in results if r[0]]
                        else:
                            query = f"SELECT * FROM {qual_tbl} WHERE `{filter_column}` = '{chosen_report}'"
                            cursor.execute(query)
                            results = cursor.fetchall()
                            col_names = [col[0] for col in cursor.description]
                            t_idx = 1 if len(col_names) > 1 else 0
                            db_columns_input = [{"db_column": str(r[t_idx]).strip(), "aliases": []} for r in results if r[t_idx]]

                    conn.close()

                    st.markdown(f"**Loaded {len(db_columns_input)} Target Columns for `{chosen_report}`:**")
                    st.markdown(" ".join([f"<span class='col-pill'>{c['db_column']}</span>" for c in db_columns_input]), unsafe_allow_html=True)
                except Exception as e:
                    st.error(f"Error loading report columns: {e}")

    else:
        sec = st.secrets.get("databricks", {})
        cat = sec.get("catalog", "workspace")
        sch = sec.get("schema", "excel_column_mapping_utility")
        default_query = f"SELECT column_name, aliases FROM {cat}.{sch}.columns_master WHERE report_name = 'Sales_Report'"
        custom_sql = st.text_area(
            "✏️ Custom SQL Query:",
            value=st.session_state.get("last_custom_meta_sql", default_query),
            help="First column returns target column names. Second column (optional) returns aliases.",
            height=100
        )
        if st.button("🚀 Execute SQL Query"):
            st.session_state["last_custom_meta_sql"] = custom_sql
            try:
                with st.spinner("Executing query on Databricks..."):
                    conn = get_databricks_connection()
                    with conn.cursor() as cursor:
                        cursor.execute(custom_sql)
                        rows = cursor.fetchall()
                        if len(cursor.description) > 1:
                            db_columns_input = [
                                {
                                    "db_column": str(r[0]).strip(),
                                    "aliases": [a.strip() for a in str(r[1] or "").split(",") if a.strip()]
                                }
                                for r in rows if r[0]
                            ]
                        else:
                            db_columns_input = [{"db_column": str(r[0]).strip(), "aliases": []} for r in rows if r[0]]
                        st.session_state["custom_sql_cols"] = db_columns_input
                    conn.close()
                    st.success(f"✓ Retrieved {len(db_columns_input)} column names.")
            except Exception as e:
                st.error(f"SQL execution error: {e}")

        if "custom_sql_cols" in st.session_state and not db_columns_input:
            db_columns_input = st.session_state["custom_sql_cols"]
            st.markdown(f"**Active Target Columns ({len(db_columns_input)}):**")
            st.markdown(" ".join([f"<span class='col-pill'>{c['db_column']}</span>" for c in db_columns_input]), unsafe_allow_html=True)

st.markdown("<div style='margin-top: 1.5rem;'></div>", unsafe_allow_html=True)

# --- 2. File Uploads (Source & Up to 5 Masters) ---
with st.container():
    col_src, col_mst = st.columns(2)
    with col_src:
        st.markdown('<div class="step-header">📂 2A. Source Spreadsheet (Required)</div>', unsafe_allow_html=True)
        excel_file = st.file_uploader("Upload incoming spreadsheet", type=["xlsx", "xls", "csv"], key="source_file")
    with col_mst:
        st.markdown('<div class="step-header">📚 2B. Master Datasets (Max 5, Optional)</div>', unsafe_allow_html=True)
        master_files = st.file_uploader(
            "Upload up to 5 master tables to join",
            type=["xlsx", "xls", "csv"],
            accept_multiple_files=True,
            key="master_files"
        )
        if master_files and len(master_files) > 5:
            st.warning("⚠️ Only the first 5 uploaded master files will be processed.")
            master_files = master_files[:5]

# --- 3. Column Matching & Review (Green/Red Visual Styler) ---
if excel_file is not None and db_columns_input:
    targets = build_targets(db_columns_input)
    available_db_columns = [t["db_column"] for t in targets]
    dropdown_options = ["-- NOT MATCHED --"] + available_db_columns

    try:
        raw_df = load_uploaded_df(excel_file)
    except Exception as e:
        st.error(f"Could not read source spreadsheet: {e}")
        st.stop()

    file_key = f"mapping_data_{excel_file.name}_{threshold}_{len(db_columns_input)}"
    if file_key not in st.session_state:
        base_results = map_columns(list(raw_df.columns), targets, threshold)
        st.session_state[file_key] = pd.DataFrame([
            {
                "Excel Column (Original)": r["excel_column"],
                "Target DB Column": r["db_column"] if r["db_column"] else "-- NOT MATCHED --",
                "Match Status": _badge(r["match_type"]),
                "Confidence": f"{round(r['score'])}%",
            }
            for r in base_results
        ])

    st.markdown("<div style='margin-top: 1.5rem;'></div>", unsafe_allow_html=True)
    st.markdown('<div class="step-header">🔍 3. Review & Validate Mappings</div>', unsafe_allow_html=True)

    edited_table = st.data_editor(
        st.session_state[file_key],
        use_container_width=True,
        hide_index=True,
        disabled=["Excel Column (Original)", "Match Status", "Confidence"],
        column_config={
            "Target DB Column": st.column_config.SelectboxColumn(
                "Target DB Column",
                options=dropdown_options,
                required=True,
                help="Select target column to link or override.",
            )
        },
        key=f"editor_{file_key}"
    )

    st.session_state[file_key] = edited_table

    reconciled_results = []
    assigned_targets = []
    for _, row in edited_table.iterrows():
        orig_col = row["Excel Column (Original)"]
        target = None if row["Target DB Column"] == "-- NOT MATCHED --" else row["Target DB Column"]
        is_manual = "Manual" in row["Match Status"] or row["Target DB Column"] != "-- NOT MATCHED --"

        reconciled_results.append({
            "excel_column": orig_col,
            "db_column": target,
            "match_type": "Manual" if is_manual else row["Match Status"],
            "score": 100.0 if target else 0.0,
        })
        if target:
            assigned_targets.append(target)

    # Real-time KPIs
    total_cols = len(reconciled_results)
    matched_count = sum(1 for r in reconciled_results if r["db_column"])
    unmatched_count = total_cols - matched_count
    match_rate = round((matched_count / total_cols * 100), 1) if total_cols else 0

    m1, m2, m3, m4 = st.columns(4)
    with m1:
        st.markdown(f'<div class="metric-card"><div class="metric-label">Source Columns</div><div class="metric-num">{total_cols}</div></div>', unsafe_allow_html=True)
    with m2:
        st.markdown(f'<div class="metric-card"><div class="metric-label">Mapped</div><div class="metric-num" style="color:#16a34a;">{matched_count}</div></div>', unsafe_allow_html=True)
    with m3:
        st.markdown(f'<div class="metric-card"><div class="metric-label">Unmapped</div><div class="metric-num" style="color:#dc2626;">{unmatched_count}</div></div>', unsafe_allow_html=True)
    with m4:
        st.markdown(f'<div class="metric-card"><div class="metric-label">Coverage</div><div class="metric-num" style="color:#2563eb;">{match_rate}%</div></div>', unsafe_allow_html=True)

    st.markdown("<div style='margin-top: 1rem;'></div>", unsafe_allow_html=True)

    # --- Green / Red Visual Status Display ---
    with st.expander("🎨 Visual Status Breakdown (Green = Mapped, Red = Unmapped)", expanded=True):
        mapped_pills = [
            f"<span class='pill-mapped'>✓ {r['excel_column']} ➔ {r['db_column']}</span>"
            for r in reconciled_results if r["db_column"]
        ]
        unmapped_pills = [
            f"<span class='pill-unmapped'>✗ {r['excel_column']}</span>"
            for r in reconciled_results if not r["db_column"]
        ]

        c_v1, c_v2 = st.columns(2)
        with c_v1:
            st.markdown(f"**🟢 Mapped Columns ({len(mapped_pills)}):**")
            if mapped_pills:
                st.markdown(" ".join(mapped_pills), unsafe_allow_html=True)
            else:
                st.caption("None mapped yet.")

        with c_v2:
            st.markdown(f"**🔴 Unmapped Columns ({len(unmapped_pills)}):**")
            if unmapped_pills:
                st.markdown(" ".join(unmapped_pills), unsafe_allow_html=True)
            else:
                st.caption("All columns mapped!")

        def style_mapping_row(row):
            is_mapped = row["Target DB Column"] != "-- NOT MATCHED --"
            bg = "background-color: #dcfce7; color: #14532d; font-weight: 500;" if is_mapped else "background-color: #ffe4e6; color: #7f1d1d; font-weight: 500;"
            return [bg] * len(row)

        styled_review = edited_table.style.apply(style_mapping_row, axis=1)
        st.markdown("<div style='margin-top: 0.8rem;'></div>", unsafe_allow_html=True)
        st.dataframe(styled_review, use_container_width=True, hide_index=True)

    duplicates = [col for col, count in Counter(assigned_targets).items() if count > 1]
    has_collision = len(duplicates) > 0
    if has_collision:
        st.error(f"⛔ **Mapping Collision:** Target column `{', '.join(duplicates)}` is mapped to multiple original columns.")

    rename_dict = {r["excel_column"]: r["db_column"] for r in reconciled_results if r["db_column"]}
    working_df = raw_df.rename(columns=rename_dict)

    # --- 4. Multi-Master Data Joiner ---
    st.markdown("<div style='margin-top: 2rem;'></div>", unsafe_allow_html=True)
    st.markdown('<div class="step-header">🔗 4. Master Data Joins (Up to 5 Masters)</div>', unsafe_allow_html=True)

    joined_df = working_df.copy()

    if master_files:
        st.caption(f"Loaded **{len(master_files)}** master dataset(s). Configure join parameters below:")
        tabs = st.tabs([f"Master {i+1}: {f.name}" for i, f in enumerate(master_files)])

        for i, (tab, mfile) in enumerate(zip(tabs, master_files)):
            with tab:
                try:
                    m_df = load_uploaded_df(mfile)
                    st.write(f"**Rows:** {m_df.shape[0]} | **Columns:** {m_df.shape[1]}")

                    c_enable, c_join = st.columns([1, 2])
                    with c_enable:
                        enable_this_join = st.checkbox(f"Enable Join for Master {i+1}", value=True, key=f"enable_m_{i}")
                    with c_join:
                        join_type = st.selectbox(
                            "Join Type",
                            ["left", "right", "inner", "outer"],
                            index=0,
                            key=f"join_type_{i}",
                            format_func=lambda x: f"{x.capitalize()} Join"
                        )

                    c_left, c_right = st.columns(2)
                    with c_left:
                        current_source_cols = list(joined_df.columns)
                        left_key = st.selectbox(
                            "Source Key Column",
                            options=current_source_cols,
                            key=f"left_key_{i}"
                        )
                    with c_right:
                        right_key = st.selectbox(
                            "Master Key Column",
                            options=list(m_df.columns),
                            key=f"right_key_{i}"
                        )

                    available_enrichment = [c for c in m_df.columns if c != right_key]
                    selected_enrichment = st.multiselect(
                        "Columns to bring from this Master:",
                        options=available_enrichment,
                        default=available_enrichment,
                        key=f"cols_m_{i}"
                    )

                    if enable_this_join and left_key and right_key:
                        master_subset = m_df[[right_key] + selected_enrichment].drop_duplicates(subset=[right_key])
                        suffix = f"_m{i+1}"
                        joined_df = pd.merge(
                            joined_df,
                            master_subset,
                            how=join_type,
                            left_on=left_key,
                            right_on=right_key,
                            suffixes=("", suffix)
                        )
                except Exception as e:
                    st.error(f"Error executing join for {mfile.name}: {e}")

        st.success(f"✓ Join pipeline complete. Resulting dimensions: {joined_df.shape[0]} rows × {joined_df.shape[1]} columns.")
    else:
        st.info("💡 Upload one or more master files in **Step 2B** to enrich your data.")

    # --- 5. Export with Color-Coded Column Selector & Preview ---
    st.markdown("<div style='margin-top: 2rem;'></div>", unsafe_allow_html=True)
    st.markdown('<div class="step-header">🚀 5. Preview & Export</div>', unsafe_allow_html=True)

    all_exportable_cols = list(joined_df.columns)
    
    # Helper sets to distinguish column provenance
    mapped_target_names = {r["db_column"] for r in reconciled_results if r["db_column"]}
    working_df_cols = set(working_df.columns)

    def get_column_type(col_name: str) -> str:
        if col_name in mapped_target_names:
            return "mapped"
        elif col_name not in working_df_cols:
            return "master"
        else:
            return "unmapped"

    def format_col_option(col_name: str) -> str:
        ctype = get_column_type(col_name)
        if ctype == "mapped":
            return f"🟢 {col_name} (Mapped)"
        elif ctype == "master":
            return f"🔵 {col_name} (Master Data)"
        else:
            return f"🔴 {col_name} (Unmapped Source)"

    with st.expander("👁️ Data Preview & Column Selection", expanded=True):
        selected_cols = st.multiselect(
            "Included Columns in Output:",
            options=all_exportable_cols,
            default=all_exportable_cols,
            format_func=format_col_option,
            help="🟢 Green: Standardized DB column | 🔴 Red: Unmapped original column | 🔵 Blue: Master lookup column"
        )

        # Render dynamic colored badges reflecting current multiselect choices
        if selected_cols:
            selected_badges = []
            for col in selected_cols:
                ctype = get_column_type(col)
                if ctype == "mapped":
                    selected_badges.append(f"<span class='pill-mapped'>✓ {col}</span>")
                elif ctype == "master":
                    selected_badges.append(f"<span class='pill-master'>+ {col}</span>")
                else:
                    selected_badges.append(f"<span class='pill-unmapped'>! {col}</span>")
            
            st.markdown(" ".join(selected_badges), unsafe_allow_html=True)
            st.write("")

            # Color-styled DataFrame preview (mapped columns green, unmapped red, master blue)
            def style_preview_column(s):
                ctype = get_column_type(s.name)
                if ctype == "mapped":
                    color = "background-color: #f0fdf4; color: #166534;"
                elif ctype == "master":
                    color = "background-color: #f0f9ff; color: #075985;"
                else:
                    color = "background-color: #fef2f2; color: #991b1b;"
                return [color] * len(s)

            styled_preview_df = joined_df[selected_cols].head(10).style.apply(style_preview_column, axis=0)
            st.dataframe(styled_preview_df, use_container_width=True)
        else:
            st.warning("Select at least one column to export.")

    filter_download = st.checkbox("Export only the selected column subset", value=False)
    final_export_df = joined_df[selected_cols] if filter_download and selected_cols else joined_df
    restrict_report = selected_cols if filter_download and selected_cols else None

    if not has_collision and len(final_export_df.columns) > 0:
        output_bytes = build_output_workbook_bytes(final_export_df, reconciled_results, restrict_report_to=restrict_report)
        out_name = f"{excel_file.name.rsplit('.', 1)[0]}_standardized.xlsx"

        st.download_button(
            label="📥 Download Standardized Spreadsheet",
            data=output_bytes,
            file_name=out_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )
    else:
        st.button("📥 Download Standardized Spreadsheet", disabled=True, use_container_width=True)

elif not db_columns_input:
    st.info("💡 Select a report or execute custom SQL in **Step 1** to populate target columns.")
else:
    st.info("💡 Upload an incoming spreadsheet in **Step 2A** to begin matching.")
